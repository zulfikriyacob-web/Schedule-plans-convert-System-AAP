import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Konfigurasi Muka Depan Web App
st.set_page_config(page_title="AAP Schedule Converter", page_icon="🏭", layout="wide")

# Set up memory (Session State) untuk butang Bypass
if 'bypass_warning' not in st.session_state:
    st.session_state.bypass_warning = False

st.markdown("""
    <style>
    .main-title {
        text-align: center;
        font-size: 3rem;
        font-weight: 800;
        background: -webkit-linear-gradient(#4facfe, #00f2fe);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 0px;
    }
    .sub-title {
        text-align: center;
        color: #888;
        font-size: 1.2rem;
        margin-bottom: 30px;
    }
    </style>
""", unsafe_allow_html=True)

st.markdown('<p class="main-title">🚀 AAP Auto-Converter</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-title">Sistem Pengekstrakan Jadual Kilang Rasmi (Logik Lanjutan)</p>', unsafe_allow_html=True)

with st.expander("💡 Cara Penggunaan & Logik Sistem (Klik Sini)"):
    st.info("""
    1. Pastikan fail Excel asal mempunyai sheet **WO LISTING DM**, **WO LISTING OH**, dan sheet **PACK**.
    2. Tarik dan lepaskan (*drag & drop*) fail tersebut ke dalam ruang muat naik di bawah.
    3. **Logik DM Aktif:** Membaca pasangan RH-L dan RH-R berdasarkan kuantiti warna.
    4. **Logik OH Aktif:** Membaca pecahan L RR, R RR, L FR, R FR mengikut kuantiti total.
    5. Sistem akan kesan Lot Number yang hilang, buang *space* tersembunyi, dan susun format *column*.
    """)

st.divider()

uploaded_file = st.file_uploader("📂 Muat naik fail Excel jadual di sini...", type=['xlsx'])

if uploaded_file is None:
    st.session_state.bypass_warning = False

def find_sheet(xls, keywords):
    for sheet in xls.sheet_names:
        if all(k.upper() in sheet.upper() for k in keywords):
            return sheet
    return None

def process_wo_sheet(df, sheet_type="DM", bypass_mode=False):
    header_idx = 1
    headers = df.iloc[header_idx].values
    df_clean = df.iloc[header_idx + 1:].copy()
    df_clean.columns = headers
    df_clean = df_clean.dropna(how='all')
    
    # Anti-Hantu: Tukar cell yang cuma ada spacebar kepada NaN
    df_clean = df_clean.replace(r'^\s*$', np.nan, regex=True)
    
    col_mapping = {
        'PROD DATE': 'PROD_DATE',
        'LOT NUMBER': 'LOT_NUMBER',
        'MODEL': 'MODEL',
        'FS CODE': 'FS_CODE',
        'COLOUR PART': 'COLOUR_PART',
        'PART NAME': 'PART_NAME',
        'Total': 'PLANNED_COLOUR_QTY',
        'WO NUM': 'WO_NUM',
        'REMARKS': 'planner_remarks',
        'WO SUPPLY TO IMC': 'WO_SUPPLY_TO_IMC_DATE',
        'Closing Date': 'DI_DATE'
    }
    
    df_clean = df_clean.rename(columns=lambda x: col_mapping.get(x, x))
    
    cols_to_ffill = ['PROD_DATE', 'LOT_NUMBER', 'MODEL', 'planner_remarks', 'WO_SUPPLY_TO_IMC_DATE', 'DI_DATE']
    
    # JIKA BYPASS: Kita jangan ffill LOT_NUMBER, biarkan ia KOSONG (Lopong) untuk baris yang salah
    if bypass_mode:
        cols_to_ffill.remove('LOT_NUMBER')
        
    for col in cols_to_ffill:
        if col in df_clean.columns:
            df_clean[col] = df_clean[col].ffill()
    
    date_columns = ['PROD_DATE', 'WO_SUPPLY_TO_IMC_DATE', 'DI_DATE']
    for col in date_columns:
        if col in df_clean.columns:
            df_clean[col] = pd.to_datetime(df_clean[col], errors='coerce')
    
    if 'DI_DATE' in df_clean.columns and df_clean['DI_DATE'].isnull().all():
        df_clean['DI_DATE'] = df_clean['PROD_DATE'] + pd.Timedelta(days=1)
        
    for col in date_columns:
        if col in df_clean.columns:
            df_clean[col] = df_clean[col].dt.date
            
    if 'COLOUR_PART' in df_clean.columns:
        df_clean['COLOUR_PART'] = df_clean['COLOUR_PART'].astype(str).str.replace('B$', '', regex=True)
        
    df_clean['CUSTOMER_TYPE'] = 'OEM'
    
    def map_colour(c):
        c = str(c)
        if 'NH547' in c: return 'BB'
        elif 'B640M' in c: return 'CRB'
        elif 'NH731P' in c: return 'CB'
        elif 'NH883P' in c: return 'PW'
        elif 'NH830' in c: return 'LS'
        elif 'NH904M' in c: return 'RG'
        elif 'NH922P' in c: return 'SD'
        elif 'R575M' in c: return 'IR'
        return ''
    
    if 'COLOUR_PART' in df_clean.columns:
        df_clean['COLOUR_CODE'] = df_clean['COLOUR_PART'].apply(map_colour)
    else:
         df_clean['COLOUR_CODE'] = ''

    # 🔴 IMPLEMENTASI LOGIC RULE KILANG BOS (DM & OH)
    def get_side_logic(name, type_sheet):
        name = str(name).upper()
        if type_sheet == "DM":
            # Logik DM: 1 Lot -> 2 Side (RH-L & RH-R)
            if 'RH-L' in name or 'LEFT' in name: return 'RH-L'
            if 'RH-R' in name or 'RIGHT' in name: return 'RH-R'
        elif type_sheet == "OH":
            # Logik OH: Pecahan Front dan Rear
            if 'L RR' in name or 'LEFT REAR' in name: return 'L RR'
            if 'R RR' in name or 'RIGHT REAR' in name: return 'R RR'
            if 'L FR' in name or 'LEFT FRONT' in name: return 'L FR'
            if 'R FR' in name or 'RIGHT FRONT' in name: return 'R FR'
        return ''
    
    if 'PART_NAME' in df_clean.columns:
        df_clean['SIDE_CODE'] = df_clean['PART_NAME'].apply(lambda x: get_side_logic(x, sheet_type))
    else:
        df_clean['SIDE_CODE'] = ''
        
    if 'WO_NUM' in df_clean.columns and 'FS_CODE' in df_clean.columns:
        df_clean['WO_NUM_STR'] = pd.to_numeric(df_clean['WO_NUM'], errors='coerce').fillna(0).astype(int).astype(str)
        df_clean['WO_NUM_STR'] = df_clean['WO_NUM_STR'].replace('0', '')
        df_clean['WO_NUM_+_FS_CODE'] = df_clean['WO_NUM_STR'] + df_clean['FS_CODE'].astype(str).fillna('')
        df_clean = df_clean.drop(columns=['WO_NUM_STR'])
    
    df_clean['LINE'] = 'L1'
    
    def split_model(m):
        m = str(m)
        if '-' in m:
            parts = m.split('-')
            return parts[0], '-'.join(parts[1:])
        return m, ''
    
    if 'MODEL' in df_clean.columns:
        df_clean[['MODEL_TYPE', 'VARIANCE']] = df_clean['MODEL'].apply(lambda x: pd.Series(split_model(x)))
    else:
        df_clean['MODEL_TYPE'] = ''
        df_clean['VARIANCE'] = ''
        
    df_clean['CAMERA_APPLICABLE'] = np.nan
    df_clean['ADDITIONAL_ATTRIBUTE'] = np.nan
    df_clean['RUN_STATUS'] = np.nan
    
    # 🔴 FORMAT LOCK: 21 Column Tetap
    target_cols = [
        'PROD_DATE', 'LOT_NUMBER', 'MODEL', 'FS_CODE', 'COLOUR_PART', 'PART_NAME', 
        'PLANNED_COLOUR_QTY', 'WO_NUM', 'planner_remarks', 'WO_SUPPLY_TO_IMC_DATE', 
        'DI_DATE', 'CUSTOMER_TYPE', 'COLOUR_CODE', 'SIDE_CODE', 'WO_NUM_+_FS_CODE', 
        'LINE', 'MODEL_TYPE', 'VARIANCE', 'CAMERA_APPLICABLE', 'ADDITIONAL_ATTRIBUTE', 'RUN_STATUS'
    ]
    
    for col in target_cols:
        if col not in df_clean.columns:
            df_clean[col] = np.nan
            
    df_final = df_clean[target_cols]
    df_final = df_final.dropna(subset=['WO_NUM'])
    
    return df_final

def process_ot_sheet(df_pack, date_col_idx, ot_col_idx, prefix):
    expected_cols = ['DATE', 'DAY', f'{prefix}_L1', f'{prefix}_L2']
    
    if ot_col_idx >= len(df_pack.columns):
        return pd.DataFrame(columns=expected_cols)
    
    df_ot = df_pack.iloc[:, [date_col_idx, ot_col_idx]].copy()
    df_ot.columns = ['DATE', f'{prefix}_L1']
    
    df_ot['DATE_PARSED'] = pd.to_datetime(df_ot['DATE'], errors='coerce')
    df_ot = df_ot.dropna(subset=['DATE_PARSED']) 
    df_ot['DATE'] = df_ot['DATE_PARSED'].dt.date
    
    df_ot[f'{prefix}_L1'] = pd.to_numeric(df_ot[f'{prefix}_L1'], errors='coerce')
    df_ot = df_ot.groupby('DATE', as_index=False).agg({f'{prefix}_L1': 'max'})
    
    df_ot['DAY'] = pd.to_datetime(df_ot['DATE']).dt.day_name().str.upper()
    df_ot[f'{prefix}_L2'] = np.nan
    
    df_ot = df_ot[expected_cols]
    return df_ot

if uploaded_file is not None:
    try:
        xls = pd.ExcelFile(uploaded_file)
        dm_sheet_name = find_sheet(xls, ['WO LISTING', 'DM'])
        oh_sheet_name = find_sheet(xls, ['WO LISTING', 'OH'])
        pack_sheet_name = find_sheet(xls, ['PACK'])
        
        if not dm_sheet_name or not oh_sheet_name or not pack_sheet_name:
            st.error("Ralat: Tidak menjumpai Sheet yang diperlukan.")
            st.stop()
            
        # 🔴 WARNING BLOCKER (Forensic Cross-Check)
        df_dm_raw_check = pd.read_excel(xls, sheet_name=dm_sheet_name, header=1)
        df_oh_raw_check = pd.read_excel(xls, sheet_name=oh_sheet_name, header=1)
        df_pack_check = pd.read_excel(xls, sheet_name=pack_sheet_name, header=None)
        
        wo_dm_lots = set(df_dm_raw_check['LOT NUMBER'].replace(r'^\s*$', np.nan, regex=True).dropna().astype(str).str.strip().unique())
        wo_oh_lots = set(df_oh_raw_check['LOT NUMBER'].replace(r'^\s*$', np.nan, regex=True).dropna().astype(str).str.strip().unique())
        all_wo_lots = wo_dm_lots.union(wo_oh_lots)
        pack_lots = set(df_pack_check.iloc[:, 2].replace(r'^\s*$', np.nan, regex=True).dropna().astype(str).str.strip().unique())
        
        missing_lots = pack_lots - all_wo_lots
        missing_lots = {lot for lot in missing_lots if lot.startswith('M')}
        
        if missing_lots and not st.session_state.bypass_warning:
            st.error("🚨 **AMARAN: DATA TERTINGGAL DALAM FAIL EXCEL ASAL!**")
            st.warning(f"Sistem mendapati Lot Number ini wujud dalam sheet **PACK**, tetapi **TIDAK DITULIS / HILANG** dalam sheet **WO LISTING DM/OH**:\n\n👉 **{', '.join(missing_lots)}**")
            st.info("💡 **TINDAKAN:** Sila pilih salah satu daripada pilihan di bawah:")
            
            col_a, col_b = st.columns(2)
            with col_a:
                if st.button("⚠️ Abaikan Amaran & Teruskan Download"):
                    st.session_state.bypass_warning = True
                    st.rerun() 
            with col_b:
                if st.button("❌ Batal & Buat Semula"):
                    st.session_state.bypass_warning = False
                    st.rerun()
            st.stop()
            
        with st.status("Kilang memproses data berjalan...", expanded=True) as status:
            st.write("Mengekstrak data barisan DM & OH dengan LOGIK KILANG...")
            df_dm_raw = pd.read_excel(xls, sheet_name=dm_sheet_name, header=None)
            df_dm = process_wo_sheet(df_dm_raw, sheet_type="DM", bypass_mode=st.session_state.bypass_warning)
            
            df_oh_raw = pd.read_excel(xls, sheet_name=oh_sheet_name, header=None)
            df_oh = process_wo_sheet(df_oh_raw, sheet_type="OH", bypass_mode=st.session_state.bypass_warning)
            
            st.write("Menyusun jadual OT (Kumpulan Tarikh Unik)...")
            df_dm_ot = process_ot_sheet(df_pack_check, 13, 15, "DM")
            df_oh_ot = process_ot_sheet(df_pack_check, 34, 36, "OH")
            
            st.write("Membungkus ke dalam format akhir...")
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_dm.to_excel(writer, sheet_name='DM', index=False)
                df_oh.to_excel(writer, sheet_name='OH', index=False)
                df_dm_ot.to_excel(writer, sheet_name='DM_OT', index=False)
                df_oh_ot.to_excel(writer, sheet_name='OH_OT', index=False)
                
                workbook = writer.book
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                                top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
                alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                for sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        for cell in row:
                            cell.border = border
                            cell.alignment = alignment
                            if isinstance(cell.value, pd.Timestamp) or type(cell.value).__name__ == 'date':
                                cell.number_format = 'DD/MM/YYYY'
                    for column_cells in ws.columns:
                        length = max(len(str(cell.value)) for cell in column_cells)
                        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 30)

            processed_data = output.getvalue()
            status.update(label="Proses Selesai Sepenuhnya!", state="complete", expanded=False)
        
        st.balloons()
        
        st.markdown("### 📊 Rumusan Ekstrak Data")
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Kuantiti WO (DM)", f"{len(df_dm)} Baris")
        col2.metric("Kuantiti WO (OH)", f"{len(df_oh)} Baris")
        col3.metric("Rekod Unik (DM_OT)", f"{len(df_dm_ot)} Hari") 
        col4.metric("Rekod Unik (OH_OT)", f"{len(df_oh_ot)} Hari")
        
        st.markdown("### 👀 Intai Jadual")
        tab1, tab2, tab3, tab4 = st.tabs(["Sheet DM", "Sheet OH", "Sheet DM_OT", "Sheet OH_OT"])
        
        with tab1:
            st.dataframe(df_dm.head(15), use_container_width=True)
        with tab2:
            st.dataframe(df_oh.head(15), use_container_width=True)
        with tab3:
            st.dataframe(df_dm_ot, use_container_width=True)
        with tab4:
            st.dataframe(df_oh_ot, use_container_width=True)
        
        st.divider()
        
        if st.session_state.bypass_warning:
            st.warning("⚠️ Fail ini dijana dengan Amaran Data Tidak Lengkap yang telah diabaikan. (Lot Number dibiarkan KOSONG)")
            
        st.download_button(
            label="📥 DOWNLOAD DATABASE SEKARANG",
            data=processed_data,
            file_name="Converted_Database_Schedule_Master.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
    except Exception as e:
        st.error(f"Maaf, ralat berlaku: {e}")
        st.stop()
